import csv
import datetime

from django.db.models import Q
from django.http import HttpResponse
from openpyxl import Workbook
from openpyxl.utils import get_column_letter


def export_csv(queryset, study_name="gtrac"):
    """
    Takes in queryset, returns csv download response
    study_name parameter is optional to control the name of the csv file.
    """
    current_date = datetime.datetime.now().strftime("%d-%b-%Y")

    response = HttpResponse(content_type="text/csv")
    response["Content-Disposition"] = 'attachment; filename="%s_samples_%s.csv"' % (
        study_name,
        current_date,
    )
    writer = csv.writer(response)
    field_names = [field.name for field in queryset.model._meta.get_fields()]
    writer.writerow(field_names)
    for row in queryset:
        values = []
        for field in field_names:
            value = getattr(row, field)
            if value is None:
                value = ""
            values.append(value)
        writer.writerow(values)
    return response


def make_naive(value):
    """
    Takes timezone aware datetime object and returns naive datetime
    """
    if value is not None:
        return value.replace(tzinfo=None)
    else:
        return None


def export_excel(queryset):
    """
    Takes queryset in and returns excel response object
    """
    response = HttpResponse(content_type="application/ms-excel")
    response[
        "Content-Disposition"
    ] = 'attachment; filename="samples_export_%s.xlsx"' % datetime.datetime.now().strftime(  # noqa E501
        "%Y-%m-%d"
    )
    workbook = Workbook()
    # Get active worksheet
    worksheet = workbook.active
    worksheet.title = "Samples"

    # Define the excel column names
    columns = [
        "Sample ID",
        "Patient ID",
        "Sample Location",
        "Sample Sublocation",
        "Sample Type",
        "Sampling Datetime",
        "Processing Datetime",
        "Frozen Datetime",
        "Sampling to Processing Time (mins)",
        "Sample Volume",
        "Sample Volume Units",
        "Freeze Thaw Count",
        "Haemolysis Reference Category (100 and above unusable)",
        "Biopsy Location",
        "Biopsy Inflamed Status",
        "Sample Comments",
        "Sample Fully Used?",
        "Created By",
        "Date Created",
        "Last Modified By",
        "Last Modified",
    ]
    row_num = 1

    # Write the column names in
    for col_num, column_title in enumerate(columns, 1):
        cell = worksheet.cell(row=row_num, column=col_num)
        cell.value = column_title
        # Setting a uniform column width
        column_letter = get_column_letter(col_num)
        column_dimensions = worksheet.column_dimensions[column_letter]
        column_dimensions.width = 20

    for sample in queryset:
        processing_time = None
        if sample.processing_datetime is not None:
            time_difference = sample.processing_datetime - sample.sample_datetime
            processing_time = int(time_difference.total_seconds() / 60)

        row_num += 1
        row = [
            sample.sample_id,
            sample.patient_id,
            sample.sample_location,
            sample.sample_sublocation,
            sample.sample_type,
            make_naive(sample.sample_datetime),
            make_naive(sample.processing_datetime),
            make_naive(sample.frozen_datetime),
            processing_time,
            sample.sample_volume,
            sample.sample_volume_units,
            sample.freeze_thaw_count,
            sample.haemolysis_reference,
            sample.biopsy_location,
            sample.biopsy_inflamed_status,
            sample.sample_comments,
            sample.is_fully_used,
            sample.created_by,
            make_naive(sample.created),
            sample.last_modified_by,
            make_naive(sample.last_modified),
        ]

        for col_num, cell_value in enumerate(row, 1):
            cell = worksheet.cell(row=row_num, column=col_num)
            cell.value = cell_value

    worksheet.freeze_panes = worksheet["A2"]
    workbook.save(response)

    return response


def queryset_by_study_name(model, study_name):
    """
    Takes the model, pass in the study_name parameter and this will
    return a filtered queryset
    """
    if study_name == "music":
        queryset = model.objects.filter(patient_id__startswith="MID-")
    elif study_name == "gidamps":
        queryset = model.objects.filter(patient_id__startswith="GID-")
    elif study_name == "minimusic":
        queryset = model.objects.filter(sample_id__startswith="MINI-")
    elif study_name == "marvel":
        queryset = model.objects.filter(
            Q(patient_id__regex=r"^[0-9]{6}$") | Q(is_marvel_study=True)
        )
    else:
        queryset = model.objects.all()
    return queryset
