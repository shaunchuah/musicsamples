import datetime

from django.contrib import messages
from django.contrib.auth import get_user_model
from django.contrib.auth.decorators import login_required
from django.core.paginator import EmptyPage, PageNotAnInteger, Paginator
from django.db.models import Count, Q
from django.db.models.functions import Trunc
from django.http import HttpResponse, JsonResponse
from django.shortcuts import get_object_or_404, redirect, render
from django.urls import reverse

# from django.views.decorators.cache import cache_page
from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from rest_framework import viewsets

from app.filters import SampleFilter
from app.forms import (
    CheckoutForm,
    DeleteForm,
    FullyUsedForm,
    ReactivateForm,
    RestoreForm,
    SampleForm,
)
from app.models import Sample
from app.serializers import (
    MultipleSampleSerializer,
    SampleExportSerializer,
    SampleIsFullyUsedSerializer,
    SampleSerializer,
)
from app.utils import export_csv, queryset_by_study_name

User = get_user_model()

SAMPLE_PAGINATION_SIZE = 100


@login_required(login_url="/login/")
def index(request):
    # Home Page
    sample_list = (
        Sample.objects.all()
        .filter(is_deleted=False)
        .filter(is_fully_used=False)
        .order_by("-sample_datetime")
    )
    sample_count = sample_list.count()
    page = request.GET.get("page", 1)
    paginator = Paginator(sample_list, SAMPLE_PAGINATION_SIZE)
    try:
        samples = paginator.page(page)
    except PageNotAnInteger:
        samples = paginator.page(1)
    except EmptyPage:
        samples = paginator.page(paginator.num_pages)
    context = {
        "sample_list": samples,
        "page_obj": samples,
        "sample_count": sample_count,
    }
    return render(request, "index.html", context)


@login_required(login_url="/login/")
def filter_by_study(request, study_name):
    # Home Page
    queryset = queryset_by_study_name(Sample, study_name)
    sample_list = (
        queryset.filter(is_deleted=False)
        .filter(is_fully_used=False)
        .order_by("-sample_datetime")
    )
    sample_count = sample_list.count()
    page = request.GET.get("page", 1)
    paginator = Paginator(sample_list, SAMPLE_PAGINATION_SIZE)
    try:
        samples = paginator.page(page)
    except PageNotAnInteger:
        samples = paginator.page(1)
    except EmptyPage:
        samples = paginator.page(paginator.num_pages)
    context = {
        "sample_list": samples,
        "page_obj": samples,
        "sample_count": sample_count,
        "study_name": study_name,
    }
    return render(request, "index.html", context)


@login_required(login_url="/login/")
def filter(request, study_name):
    queryset = queryset_by_study_name(Sample, study_name)

    # Perform filtering
    sample_filter = SampleFilter(request.GET, queryset=queryset)
    sample_list = sample_filter.qs
    sample_count = sample_list.count()

    # Pagination
    page = request.GET.get("page", 1)
    paginator = Paginator(sample_list, SAMPLE_PAGINATION_SIZE)
    try:
        samples = paginator.page(page)
    except PageNotAnInteger:
        samples = paginator.page(1)
    except EmptyPage:
        samples = paginator.page(paginator.num_pages)

    # Allows paginator to reconstruct the initial query string
    parameter_string = ""
    for i in request.GET:
        if i != "page":
            val = request.GET.get(i)
            parameter_string += f"&{i}={val}"

    context = {
        "sample_list": samples,
        "sample_count": sample_count,
        "sample_filter": sample_filter,
        "parameter_string": parameter_string,
        "study_name": study_name,
    }
    return render(request, "filter.html", context)


@login_required(login_url="/login/")
def filter_export_csv(request, study_name):
    queryset = queryset_by_study_name(Sample, study_name)
    sample_filter = SampleFilter(request.GET, queryset=queryset)
    sample_list = sample_filter.qs
    return export_csv(sample_list)


@login_required(login_url="/login/")
def used_samples(request):
    sample_list = (
        Sample.objects.all()
        .filter(is_deleted=False)
        .filter(is_fully_used=True)
        .order_by("-last_modified")
    )
    sample_count = sample_list.count()
    page = request.GET.get("page", 1)
    paginator = Paginator(sample_list, SAMPLE_PAGINATION_SIZE)
    try:
        samples = paginator.page(page)
    except PageNotAnInteger:
        samples = paginator.page(1)
    except EmptyPage:
        samples = paginator.page(paginator.num_pages)
    context = {
        "sample_list": samples,
        "page_obj": samples,
        "sample_count": sample_count,
    }
    return render(request, "samples/used_samples.html", context)


@login_required(login_url="/login/")
def used_samples_search(request):
    query_string = ""
    if ("q" in request.GET) and request.GET["q"].strip():
        query_string = request.GET.get("q")
        sample_list = (
            Sample.objects.filter(
                Q(sample_id__icontains=query_string)
                | Q(patient_id__icontains=query_string)
                | Q(sample_location__icontains=query_string)
                | Q(sample_sublocation__icontains=query_string)
                | Q(sample_type__icontains=query_string)
                | Q(sample_comments__icontains=query_string)
            )
            .filter(is_fully_used=True)
            .filter(is_deleted=False)
        )
        sample_count = sample_list.count()
        return render(
            request,
            "samples/used_samples.html",
            {
                "query_string": query_string,
                "sample_list": sample_list,
                "sample_count": sample_count,
            },
        )
    else:
        return render(
            request,
            "samples/used_samples.html",
            {"query_string": "Null", "sample_count": 0},
        )


@login_required(login_url="/login/")
def used_samples_archive_all(request):
    """
    This function will retrieve all used samples and remove their last location.
    This helps to keep the database clean.
    """
    sample_list = (
        Sample.objects.filter(is_deleted=False)
        .filter(is_fully_used=True)
        .exclude(sample_location="used")
    )

    number_of_samples = sample_list.count()
    if (
        number_of_samples == 0
    ):  # if no samples need updating; skip the database update step
        messages.error(request, "No samples to update.")
    else:
        for sample in sample_list:
            sample.sample_location = "used"
            sample.sample_sublocation = ""
            sample.save()
        messages.success(
            request,
            "Used samples locations archived successfully."
            f" ({number_of_samples} samples updated)",
        )
    return redirect(reverse("used_samples"))


@login_required(login_url="/login/")
# @cache_page(60 * 10)  # Cache page for 10 minutes
def analytics(request):
    # Analytics Page

    # Get last 12 months date
    current_date = datetime.datetime.now(datetime.timezone.utc)
    months_ago = 12
    twelve_months_previous_date = current_date - datetime.timedelta(
        days=(months_ago * 365 / 12)
    )

    total_samples = Sample.objects.all().filter(is_deleted=False).count()
    total_active_samples = (
        Sample.objects.all()
        .filter(is_deleted=False)
        .filter(is_fully_used=False)
        .count()
    )
    samples_by_month = (
        Sample.objects.all()
        .filter(is_deleted=False, sample_datetime__gte=twelve_months_previous_date)
        .order_by()
        .annotate(sample_month=Trunc("sample_datetime", "month"))
        .values("sample_month")
        .annotate(sample_count=Count("id"))
        .order_by("sample_month")
    )
    samples_by_type = (
        Sample.objects.all()
        .filter(is_deleted=False)
        .filter(is_fully_used=False)
        .order_by()
        .values("sample_type")
        .annotate(sample_type_count=Count("id"))
    )
    samples_by_location = (
        Sample.objects.all()
        .filter(is_deleted=False)
        .filter(is_fully_used=False)
        .order_by()
        .values("sample_location")
        .annotate(sample_location_count=Count("id"))
    )
    context = {
        "total_samples": total_samples,
        "samples_by_month": samples_by_month,
        "samples_by_type": samples_by_type,
        "samples_by_location": samples_by_location,
        "total_active_samples": total_active_samples,
    }
    return render(request, "analytics.html", context)


@login_required(login_url="/login/")
# @cache_page(3 * 60)  # Cache page for 3 minutes
def gid_overview(request):
    # Analytics --> Sample Overview Table Page
    sample_categories = (
        Sample.objects.filter(is_deleted=False)
        .filter(is_fully_used=False)
        .order_by()
        .values("sample_type")
        .distinct()
    )
    sample_category_list = []
    for item in sample_categories:
        sample_category_list.append(item["sample_type"])
    if "q" in request.GET:
        query = request.GET.get("q")
        sample_list = (
            Sample.objects.filter(is_deleted=False)
            .filter(is_fully_used=False)
            .filter(sample_type=query)
            .order_by("patient_id")
        )
    else:
        query = None
        sample_list = Sample.objects.none()
    context = {
        "sample_list": sample_list,
        "sample_category_list": sample_category_list,
        "query": query,
    }
    return render(request, "gid_overview.html", context)


@login_required(login_url="/login/")
def reference(request):
    # Reference static page for publishing lab protocols
    return render(request, "reference.html")


@login_required(login_url="/login/")
def account(request):
    # User account page showing last 20 recently accessed samples
    sample_list = (
        Sample.objects.all()
        .filter(is_deleted=False)
        .filter(last_modified_by=request.user.username)
        .order_by("-last_modified")[:20]
    )
    context = {"sample_list": sample_list}
    return render(request, "account.html", context)


@login_required(login_url="/login/")
def data_export(request):
    return render(request, "data_export.html")


@login_required(login_url="/login/")
def sample_archive(request):
    # Deleted samples page for samples which have been soft deleted
    sample_list = (
        Sample.objects.all().filter(is_deleted=True).order_by("-last_modified")
    )
    context = {"sample_list": sample_list}
    return render(request, "samples/sample-archive.html", context)


@login_required(login_url="/login/")
def sample_add(request):
    # Add mew sample page
    if request.method == "POST":
        form = SampleForm(request.POST)
        if form.is_valid():
            sample = form.save(commit=False)
            sample.created_by = request.user.username
            sample.last_modified_by = request.user.username
            sample.save()
            messages.success(request, "Sample registered successfully.")
            return redirect(reverse("sample_add"))
        else:
            messages.error(request, "Form is not valid.")
    else:
        form = SampleForm()
    return render(
        request,
        "samples/sample-add.html",
        {"form": form, "page_title": "Add New Sample"},
    )


def historical_changes(query):
    # Historical changes integrates simple history into the sample detail page
    changes = []
    if query is not None:
        last = query.first()
        for all_changes in range(query.count()):
            new_record, old_record = last, last.prev_record
            if old_record is not None:
                delta = new_record.diff_against(old_record)
                changes.append(delta)
                last = old_record
        return changes


@login_required(login_url="/login/")
def sample_detail(request, pk):
    # retrieves sample history and also linked notes both public and private
    sample = get_object_or_404(Sample, pk=pk)
    sample_history = sample.history.filter(id=pk)
    changes = historical_changes(sample_history)
    first_change = sample_history.first()
    processing_time = None
    if sample.processing_datetime is not None:
        time_difference = sample.processing_datetime - sample.sample_datetime
        processing_time = int(time_difference.total_seconds() / 60)
    return render(
        request,
        "samples/sample-detail.html",
        {
            "sample": sample,
            "changes": changes,
            "first": first_change,
            "processing_time": processing_time,
        },
    )


@login_required(login_url="/login/")
def sample_edit(request, pk):
    # Sample Edit Page
    sample = get_object_or_404(Sample, pk=pk)
    if request.method == "POST":
        form = SampleForm(request.POST, instance=sample)
        if form.is_valid():
            sample = form.save(commit=False)
            sample.last_modified_by = request.user.username
            sample.save()
            messages.success(request, "Sample updated successfully.")
            next_url = request.GET.get("next")
            if next_url:
                return redirect(next_url)
            else:
                return redirect("/")
    else:
        form = SampleForm(instance=sample)
    return render(
        request,
        "samples/sample-add.html",
        {"form": form, "page_title": "Update Sample"},
    )


@login_required(login_url="/login/")
def sample_search(request):
    # Sample search in home page
    query_string = ""
    if ("q" in request.GET) and request.GET["q"].strip():
        query_string = request.GET.get("q")
        sample_list = (
            Sample.objects.filter(
                Q(sample_id__icontains=query_string)
                | Q(patient_id__icontains=query_string)
                | Q(sample_location__icontains=query_string)
                | Q(sample_sublocation__icontains=query_string)
                | Q(sample_type__icontains=query_string)
                | Q(sample_comments__icontains=query_string)
            )
            .filter(is_fully_used=False)
            .filter(is_deleted=False)
        )
        sample_count = sample_list.count()
        return render(
            request,
            "index.html",
            {
                "query_string": query_string,
                "sample_list": sample_list,
                "sample_count": sample_count,
            },
        )
    else:
        return render(
            request, "index.html", {"query_string": "Null", "sample_count": 0}
        )


@login_required(login_url="/login/")
def sample_checkout(request, pk):
    # Sample Checkout - Quick update of sample location from home page
    sample = get_object_or_404(Sample, pk=pk)
    if request.method == "POST":
        form = CheckoutForm(request.POST, instance=sample)
        if form.is_valid():
            sample = form.save(commit=False)
            sample.last_modified_by = request.user.username
            sample.save()
            messages.success(request, "Sample updated successfully.")
            next_url = request.GET.get("next")
            if next_url:
                return redirect(next_url)
            else:
                return redirect("/")
    else:
        form = CheckoutForm(instance=sample)
    return render(request, "samples/sample-checkout.html", {"form": form})


@login_required(login_url="/login/")
def sample_delete(request, pk):
    # Soft deletion method for samples
    sample = get_object_or_404(Sample, pk=pk)
    if request.method == "POST":
        form = DeleteForm(request.POST, instance=sample)
        if form.is_valid():
            sample = form.save(commit=False)
            sample.last_modified_by = request.user.username
            sample.save()
            messages.success(request, "Sample deleted.")
            next_url = request.GET.get("next")
            if next_url:
                return redirect(next_url)
            else:
                return redirect("/")
    else:
        form = DeleteForm(instance=sample)
    return render(request, "samples/sample-delete.html", {"form": form})


@login_required(login_url="/login/")
def sample_restore(request, pk):
    # Restore soft-deleted sample
    sample = get_object_or_404(Sample, pk=pk)
    if request.method == "POST":
        form = RestoreForm(request.POST, instance=sample)
        if form.is_valid():
            sample = form.save(commit=False)
            sample.last_modified_by = request.user.username
            sample.save()
            messages.success(request, "Sample restored.")
            next_url = request.GET.get("next")
            if next_url:
                return redirect(next_url)
            else:
                return redirect("/")
    else:
        form = RestoreForm(instance=sample)
    return render(request, "samples/sample-restore.html", {"form": form})


@login_required(login_url="/login/")
def sample_fully_used(request, pk):
    # Mark sample as fully used
    sample = get_object_or_404(Sample, pk=pk)
    if request.method == "POST":
        form = FullyUsedForm(request.POST, instance=sample)
        if form.is_valid():
            sample = form.save(commit=False)
            sample.last_modified_by = request.user.username
            sample.save()
            messages.success(request, "Sample marked as fully used.")
            next_url = request.GET.get("next")
            if next_url:
                return redirect(next_url)
            else:
                return redirect("/")
    else:
        form = FullyUsedForm(instance=sample)
    return render(request, "samples/sample-fullyused.html", {"form": form})


@login_required(login_url="/login/")
def reactivate_sample(request, pk):
    # Reactive sample which has been marked as fully used
    sample = get_object_or_404(Sample, pk=pk)
    if request.method == "POST":
        form = ReactivateForm(request.POST, instance=sample)
        if form.is_valid():
            sample = form.save(commit=False)
            sample.last_modified_by = request.user.username
            sample.save()
            messages.success(request, "Sample restored.")
            next_url = request.GET.get("next")
            if next_url:
                return redirect(next_url)
            else:
                return redirect("/")
    else:
        form = ReactivateForm(instance=sample)
    return render(request, "samples/sample-reactivate.html", {"form": form})


@login_required(login_url="/login/")
def export_excel(request, study_name):
    queryset = queryset_by_study_name(Sample, study_name)

    # Exports custom views based on the search string otherwise exports entire database
    def make_naive(value):
        if value is not None:
            return value.replace(tzinfo=None)
        else:
            return None

    query_string = ""
    if ("q" in request.GET) and request.GET["q"].strip():
        query_string = request.GET.get("q")
        samples_queryset = queryset.filter(
            Q(sample_id__icontains=query_string)
            | Q(patient_id__icontains=query_string)
            | Q(sample_location__icontains=query_string)
            | Q(sample_type__icontains=query_string)
            | Q(sample_comments__icontains=query_string)
        )
    else:
        samples_queryset = queryset.filter(is_deleted=False)

    response = HttpResponse(content_type="application/ms-excel")
    response[
        "Content-Disposition"
    ] = 'attachment; filename="samples_export_%s.xlsx"' % datetime.datetime.now().strftime(  # noqa E501
        "%Y-%m-%d"
    )

    workbook = Workbook()
    # Get active worksheet
    worksheet = workbook.active
    worksheet.title = "Samples"

    # Define the excel column names
    columns = [
        "Sample ID",
        "Patient ID",
        "Sample Location",
        "Sample Sublocation",
        "Sample Type",
        "Sampling Datetime",
        "Processing Datetime",
        "Frozen Datetime",
        "Sampling to Processing Time (mins)",
        "Sample Volume",
        "Sample Volume Units",
        "Freeze Thaw Count",
        "Haemolysis Reference Category (100 and above unusable)",
        "Biopsy Location",
        "Biopsy Inflamed Status",
        "Sample Comments",
        "Sample Fully Used?",
        "Created By",
        "Date Created",
        "Last Modified By",
        "Last Modified",
    ]
    row_num = 1

    # Write the column names in
    for col_num, column_title in enumerate(columns, 1):
        cell = worksheet.cell(row=row_num, column=col_num)
        cell.value = column_title
        # Setting a uniform column width
        column_letter = get_column_letter(col_num)
        column_dimensions = worksheet.column_dimensions[column_letter]
        column_dimensions.width = 20

    for sample in samples_queryset:
        processing_time = None
        if sample.processing_datetime is not None:
            time_difference = sample.processing_datetime - sample.sample_datetime
            processing_time = int(time_difference.total_seconds() / 60)

        row_num += 1
        row = [
            sample.sample_id,
            sample.patient_id,
            sample.sample_location,
            sample.sample_sublocation,
            sample.sample_type,
            make_naive(sample.sample_datetime),
            make_naive(sample.processing_datetime),
            make_naive(sample.frozen_datetime),
            processing_time,
            sample.sample_volume,
            sample.sample_volume_units,
            sample.freeze_thaw_count,
            sample.haemolysis_reference,
            sample.biopsy_location,
            sample.biopsy_inflamed_status,
            sample.sample_comments,
            sample.is_fully_used,
            sample.created_by,
            make_naive(sample.created),
            sample.last_modified_by,
            make_naive(sample.last_modified),
        ]

        for col_num, cell_value in enumerate(row, 1):
            cell = worksheet.cell(row=row_num, column=col_num)
            cell.value = cell_value

    worksheet.freeze_panes = worksheet["A2"]
    workbook.save(response)
    return response


@login_required(login_url="/login/")
def export_study_samples(request, study_name):
    """
    Takes a study name for export. ?study_name=music
    Options:
        music: patient_id starts with MID
        music_edinburgh: patient_id starts with MID-90
        music_glasgow: patient_id starts with MID-91
        gidamps: patient_id starts with GID
        marvel: is_marvel_study=True
    """

    if study_name == "music":
        queryset = Sample.objects.filter(patient_id__startswith="MID")
    elif study_name == "music_edinburgh":
        queryset = Sample.objects.filter(patient_id__startswith="MID-90")
    elif study_name == "music_glasgow":
        queryset = Sample.objects.filter(patient_id__startswith="MID-91")
    elif study_name == "gidamps":
        queryset = Sample.objects.filter(patient_id__startswith="GID")
    elif study_name == "marvel":
        queryset = Sample.objects.filter(is_marvel_study=True)

    return export_csv(queryset, study_name)


#############################################################################
# AUTOCOMPLETE/AJAX SECTION #################################################
#############################################################################


@login_required(login_url="/login/")
def autocomplete_locations(request):
    # Helps speed up sample adding by autocompleting already existing locations
    if "term" in request.GET:
        qs = (
            Sample.objects.filter(is_deleted=False)
            .filter(sample_location__icontains=request.GET.get("term"))
            .order_by()
            .values("sample_location")
            .distinct()
        )
    else:
        qs = (
            Sample.objects.filter(is_deleted=False)
            .order_by()
            .values("sample_location")
            .distinct()
        )
    locations = list()
    for sample in qs:
        locations.append(sample["sample_location"])
    return JsonResponse(locations, safe=False)


@login_required(login_url="/login/")
def autocomplete_patient_id(request):
    # Helps speed up sample adding by locating existing patient IDs
    if "term" in request.GET:
        qs = (
            Sample.objects.filter(is_deleted=False)
            .filter(patient_id__icontains=request.GET.get("term"))
            .order_by()
            .values("patient_id")
            .distinct()
        )
    else:
        qs = (
            Sample.objects.filter(is_deleted=False)
            .order_by()
            .values("patient_id")
            .distinct()
        )
    patients = list()
    for sample in qs:
        patients.append(sample["patient_id"])
    return JsonResponse(patients, safe=False)


##############################################################################
# REST FRAMEWORK/BARCODE APIS ################################################
##############################################################################
# The bulk QR scanning capability is dependent on using JavaScript - jQuery in
# this instance - in the frontend and Django Rest Framework in the backend.
#
# Barcode readers typically terminate the scan with either the 'enter' or
# 'tab' key.
#
# The barcode scanning page listens for both enter and tab keys
#
# When these keys are heard an ajax query is sent to the API /api/<Sample ID>/
# This attempts to match the scanned label against an existing sample and
# if so updates the sample's location.
# If it does not succeed it returns an error response - Object not found.
# A warning message is then displayed.
#
# This allows you to install physical hardpoints - eg arrival and departure
# from sites and if you scan every sample that goes through you can begin to
# track all their locations.


class SampleViewSet(viewsets.ModelViewSet):
    """
    API endpoint that allows samples to be viewed and edited
    Lookup field set to the barcode ID instead of the default Django
    autoincrementing id system
    """

    queryset = Sample.objects.filter(is_deleted=False)
    serializer_class = SampleSerializer
    lookup_field = "sample_id"
    filterset_fields = ["sample_type"]

    def perform_update(self, serializer):
        serializer.save(
            last_modified_by=self.request.user.username,
        )


class SampleIsFullyUsedViewSet(viewsets.ModelViewSet):
    queryset = Sample.objects.filter(is_deleted=False)
    serializer_class = SampleIsFullyUsedSerializer
    lookup_field = "sample_id"

    def perform_update(self, serializer):
        serializer.save(
            last_modified_by=self.request.user.username,
        )


class MultipleSampleViewSet(viewsets.ModelViewSet):
    queryset = Sample.objects.filter(is_deleted=False)
    serializer_class = MultipleSampleSerializer
    lookup_field = "sample_id"

    def perform_create(self, serializer):
        serializer.save(
            created_by=self.request.user.username,
            last_modified_by=self.request.user.username,
        )


class SampleExportViewset(viewsets.ReadOnlyModelViewSet):
    queryset = Sample.objects.filter(is_deleted=False).filter(
        patient_id__startswith="GID"
    )
    serializer_class = SampleExportSerializer
    lookup_field = "sample_id"


class AllSampleExportViewset(viewsets.ReadOnlyModelViewSet):
    queryset = Sample.objects.filter(is_deleted=False)
    serializer_class = SampleExportSerializer
    lookup_field = "sample_id"

    def get_queryset(self):
        queryset = self.queryset
        patient_id_starts_with = self.request.query_params.get("patient_id_starts_with")
        if patient_id_starts_with is not None:
            queryset = queryset.filter(patient_id__istartswith=patient_id_starts_with)
        return queryset


@login_required(login_url="/login/")
def barcode(request):
    # Returns the page - frontend logic is in the barcode.html template file
    return render(request, "barcode.html")


@login_required(login_url="/login/")
def barcode_samples_used(request):
    return render(request, "barcode-markused.html")


@login_required(login_url="/login/")
def barcode_add_multiple(request):
    return render(request, "barcode-addmultiple.html")
